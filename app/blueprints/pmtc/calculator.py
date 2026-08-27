"""
calculator.py -- Calculations engine skeleton.
Pattern: copy master workbook to temp, write inputs, LibreOffice recalc, read outputs.
All reads/writes via named ranges -- never hardcoded cell addresses.
"""
import os, shutil, subprocess, tempfile, platform
from pathlib import Path
import openpyxl

# Path to master workbook (never modified -- all writes go to temp copies)
MASTER_WORKBOOK = Path(__file__).parent.parent.parent.parent / 'WORKBOOK.xlsx'

# ---------------------------------------------------------------------------
# Discovery sheet assumption overrides
# Map: session key -> (row, col, is_pct)
# Fill this from the actual workbook's Discovery sheet before Phase 3 build.
# ---------------------------------------------------------------------------
_DISC_MAP = {
    # 'session_key': (row, col, is_pct),
    # Example: 'it_labor_rate': (15, 3, False),
}

_PCT_KEYS = set()  # keys where value needs /100 conversion (stored as decimal in workbook)

# ---------------------------------------------------------------------------
# Named range -> (sheet, cell) helper
# ---------------------------------------------------------------------------
def _nr(wb, name):
    dest = list(wb.defined_names[name].destinations)
    if not dest:
        raise KeyError(f'Named range not found: {name}')
    sheet_title, coord = dest[0]
    return wb[sheet_title][coord]

# ---------------------------------------------------------------------------
# LibreOffice binary detection
# ---------------------------------------------------------------------------
def _lo_binary():
    if platform.system() == 'Windows':
        for candidate in [
            r'C:\Program Files\LibreOffice\program\soffice.exe',
            r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        ]:
            if Path(candidate).exists():
                return candidate
        raise FileNotFoundError('soffice.exe not found -- install LibreOffice')
    return shutil.which('libreoffice') or shutil.which('soffice') or 'libreoffice'

# ---------------------------------------------------------------------------
# Read workbook defaults at session start
# ---------------------------------------------------------------------------
def read_defaults():
    """
    Read assumption and investment defaults from the master workbook.
    Called once at session start; stored in session['assumption_defaults'].
    Returns dict of {field_name: value}.
    """
    wb = openpyxl.load_workbook(MASTER_WORKBOOK, data_only=True)
    defaults = {}
    # TODO: add named range reads for each assumption field
    # Example: defaults['it_labor_rate'] = _nr(wb, 'LaborRate').value or 87
    return defaults

# ---------------------------------------------------------------------------
# Main calculation
# ---------------------------------------------------------------------------
def run_calculation(profile, priorities, assumptions=None):
    """
    Write profile inputs + assumption overrides to temp workbook,
    trigger LibreOffice recalculation, read output KPIs.
    Returns dict of KPI values.
    """
    tmp_dir = tempfile.mkdtemp()
    tmp_wb_path = Path(tmp_dir) / 'workbook.xlsx'
    shutil.copy(MASTER_WORKBOOK, tmp_wb_path)

    wb = openpyxl.load_workbook(tmp_wb_path)

    # --- Write profile inputs via named ranges ---
    # TODO: fill in named range writes from actual workbook audit
    # Example:
    # _nr(wb, 'CompanyName').value = profile.get('company', '')
    # _nr(wb, 'AnnualRevenue').value = profile.get('revenue', 500)
    # for i in range(1, 8):
    #     _nr(wb, f'Ch{i}Priority').value = priorities.get(f'ch{i}', 'None')

    # --- Write assumption overrides ---
    if assumptions and _DISC_MAP:
        ws_disc = wb['Discovery']  # update sheet name to match actual workbook
        for key, (row, col, is_pct) in _DISC_MAP.items():
            if key in assumptions:
                raw = assumptions[key]
                val = float(str(raw).replace(',', ''))
                ws_disc.cell(row=row, column=col).value = val / 100 if is_pct else val

    wb.save(tmp_wb_path)

    # --- LibreOffice recalculation ---
    lo = _lo_binary()
    result = subprocess.run(
        [lo, '--headless', '--nofirststartwizard',
         '--convert-to', 'xlsx', '--outdir', tmp_dir,
         tmp_wb_path.as_uri()],
        capture_output=True, timeout=120
    )
    if result.returncode != 0:
        shutil.rmtree(tmp_dir)
        raise RuntimeError(f'LibreOffice recalc failed: {result.stderr.decode()}')

    # --- Read outputs ---
    wb2 = openpyxl.load_workbook(tmp_wb_path, data_only=True)
    kpis = _read_outputs(wb2)
    shutil.rmtree(tmp_dir)

    return kpis

def _read_outputs(wb):
    """Read all KPI output values from the recalculated workbook."""
    result = {}
    # TODO: fill in named range reads from actual workbook audit
    # Example:
    # result['roi'] = _nr(wb, 'ROI').value or 0
    # result['payback_mo'] = _nr(wb, 'PaybackMonths').value or 0
    # result['benefit_3y'] = _nr(wb, 'Benefits3Yr').value or 0
    # result['npv'] = _nr(wb, 'NPV').value or 0
    # result['irr'] = _nr(wb, 'IRR').value or 0
    # result['benefit_ann'] = _nr(wb, 'AvgAnnualBenefit').value or 0
    # result['codn_mo'] = _nr(wb, 'CODNMonthly').value or 0
    # result['fte_avg_annual'] = _nr(wb, 'FTESaved').value or 0
    return result
